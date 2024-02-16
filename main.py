import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook('sample.xlsx')
sheet = wb['Sheet1']
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    new_price = cell.value * 0.9
    price_cell = sheet.cell(row, 5)
    price_cell.value = new_price
wb.save('new_edited.xlsx')
